# Copyright (c) 2026, Miyano Việt Nam.
"""Xuất bảng chấm công tháng ra Excel **có màu** — file trông như bảng trên màn hình.

Nút Export mặc định của query report đi qua `frappe.desk.query_report.export_query`, dựng file bằng
`make_xlsx()` từ `columns` + `result`. Đường đó không có chỗ móc để tô màu: mọi ô ra file đều trắng
trơn. Module này là đường xuất riêng của báo cáo `Monthly Attendance Report`, dựng workbook bằng
openpyxl nên kiểm soát được từng ô.

Hai đường ra:

- `build_workbook(columns, data, filters)` — thuần hàm, không đụng HTTP, để test soi từng ô.
- `download(filters, visible_idx)` — endpoint cho nút Export trên report.

Màu **không định nghĩa lại ở đây**: lấy nguyên `STATE_STYLE` của `monthly_attendance_report`, dùng
cặp màu nền sáng (`bg`/`fg`) — bản nền tối chỉ có nghĩa trong Desk. State của từng ô cũng không tính
lại: `execute()` đã gắn sẵn `_state_<day>` lên mỗi dòng, y như formatter JS đọc.

CHỈ ĐỌC: không ghi Attendance, không đụng `status`/`leave_type`/`half_day_status` → lương bất biến.
"""

from io import BytesIO
from math import ceil

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import frappe
from frappe import _
from frappe.utils import cint, flt

from hrms.hr.attendance_legend import legend_pairs
from hrms.hr.report.monthly_attendance_report.monthly_attendance_report import (
	STATE_STYLE,
	day_state,
	execute,
	get_code_map,
	weekday_label,
)
from hrms.miyano_xlsx import (
	BORDER,
	CENTER,
	HEADER_ALIGN,
	HEADER_BG,
	HEADER_FG,
	HEADER_HEIGHT,
	LEFT,
	excel_width,
	period_line,
	write_letterhead,
	write_signatures,
)

# Chú thích tối đa 10 dòng; quá thì tràn sang cụm cột kế bên, không kéo dài xuống dưới.
MAX_LEGEND_ROWS = 10

# Chữ "Chú thích" — hằng số thuần, KHÔNG bọc `_()`: đây là site tiếng Việt, chuỗi nguồn đã là bản
# hiển thị, và bọc `_()` ở module level sẽ đóng băng bản dịch ngay lúc import.
LEGEND_LABEL = "Chú thích"

# Chữ "Chú thích" nằm ở cột Nhân viên, không phải cột STT: cột STT hẹp, chữ tràn ra ngoài trông
# như dữ liệu lạc dòng.
LEGEND_LABEL_COLUMN = 2

# Bề rộng (tính bằng ký tự) dành cho phần nghĩa của ký hiệu — đủ cho nhãn dài nhất trong
# `Attendance Code` ("Ngày nghỉ / ngoài thời gian làm việc").
LEGEND_MEANING_CHARS = 34

# Hai cột đầu (STT, Nhân viên) mang dữ liệu văn bản; khối chú thích bắt đầu từ cột ngày thứ nhất
# vì cột ngày hẹp, vừa đúng một ô ký hiệu.
FIRST_DAY_COLUMN = 3

# Bảng bắt đầu sau khối tiêu đề thư (6 dòng chữ + 1 dòng trống). Tiêu đề bảng chiếm hai dòng
# (số ngày, rồi thứ), nên dữ liệu bắt đầu ngay dưới.
HEADER_ROW = 8
WEEKDAY_ROW = HEADER_ROW + 1
FIRST_DATA_ROW = WEEKDAY_ROW + 1

# Cột STT thay cột "Mã NV" của báo cáo: bản in cần số thứ tự, mã hệ thống (HR-EMP-00001) chỉ tốn
# chỗ. Không đổi `columns` của report vì trên màn hình cột Mã NV là liên kết bấm sang hồ sơ.
STT_COLUMN = {"fieldname": "stt", "label": "STT", "fieldtype": "Int", "width": 44}

PLAIN_FONT = Font()
# Cột số nào cần khác thường: Tổng công là cột chủ đạo nên in đậm; STT chỉ để đánh dòng nên làm mờ
# đi, không tranh mắt với số liệu.
NUMBER_FONTS = {"tong_cong": Font(bold=True), "stt": Font(color="FF6B7280")}

# Bề rộng tối thiểu của cột không phải cột ngày, để nhãn xuống dòng chứ không bị cắt cụt.
MIN_TEXT_WIDTH = 11.0

# Bề rộng cột ngày: vừa đủ mã dài nhất (`1/2P`), giữ cả tháng lọt một trang ngang khi in.
DAY_WIDTH = 5.0


def legend_layout(count: int, max_rows: int = MAX_LEGEND_ROWS) -> tuple[int, int]:
	"""(số cụm cột, số dòng) để xếp `count` ký hiệu mà không vượt `max_rows` dòng.

	Chia đều cho các cụm chứ không nhồi đầy cụm đầu rồi bỏ cụm cuối lơ thơ: 21 mã ra 3 cụm x 7
	dòng, không phải 10 + 10 + 1."""
	if count <= 0:
		return 0, 0
	groups = ceil(count / max_rows)
	return groups, ceil(count / groups)


def fill_for(state: str) -> tuple[PatternFill, Font] | tuple[None, None]:
	"""(nền, chữ) của một state màu — dùng chung cặp màu nền sáng với report và bản in."""
	style = STATE_STYLE.get(state or "")
	if not style:
		return None, None
	return (
		PatternFill("solid", start_color="FF" + style["bg"].lstrip("#").upper()),
		Font(bold=True, color="FF" + style["fg"].lstrip("#").upper()),
	)


def build_workbook(
	columns: list[dict],
	data: list[dict],
	filters: dict | None = None,
	signatures: dict | None = None,
) -> Workbook:
	"""Dựng workbook đã tô màu từ đúng `columns`/`data` mà `execute()` của report trả về.

	`signatures` là hai cái tên trên khối trình ký (`prepared_by` / `approved_by`) — người xuất
	file điền ở hộp thoại Export; để trống thì chỉ in chức danh, ký tên bằng tay."""
	filters = frappe._dict(filters or {})
	rows = [r for r in data if r and r.get("employee")]
	columns = excel_columns(columns)

	wb = Workbook()
	ws = wb.active
	ws.title = _("Chấm công")

	last_col = len(columns)
	write_titles(ws, filters, last_col)
	write_header(ws, columns, filters)
	end_row = write_rows(ws, columns, rows, FIRST_DATA_ROW)
	set_widths(ws, columns)

	# đóng băng dưới tiêu đề, bên phải cột tên: cuộn kiểu gì cũng còn biết đang xem ai, ngày nào
	ws.freeze_panes = f"{get_column_letter(FIRST_DAY_COLUMN)}{FIRST_DATA_ROW}"

	legend_end = write_legend(ws, columns, end_row + 2)
	write_signatures(ws, last_col, legend_end + 2, signatures, filters.get("company"), FIRST_DAY_COLUMN)
	setup_print(ws, last_col)
	return wb


def excel_columns(columns: list[dict]) -> list[dict]:
	"""Cột của file: thay "Mã NV" bằng STT, giữ nguyên phần còn lại của báo cáo."""
	return [STT_COLUMN if c.get("fieldname") == "employee" else c for c in columns]


def write_titles(ws, filters: dict, last_col: int) -> None:
	"""Tiêu đề thư dùng chung với các biểu mẫu khác (`hrms/miyano_xlsx.py`)."""
	# neo theo `HEADER_ROW` chứ không theo số dòng pháp nhân vừa ghi: thiếu MST hay địa chỉ thì
	# khối trên ngắn lại, nhưng bảng vẫn phải bắt đầu đúng chỗ mọi nơi khác trong module trông đợi.
	write_letterhead(
		ws,
		filters.get("company"),
		last_col,
		_("BẢNG CHẤM CÔNG"),
		period_line(filters),
		title_row=HEADER_ROW - 3,  # tên bảng, kỳ công, rồi một dòng trống trước tiêu đề bảng
	)


def write_header(ws, columns: list[dict], filters: dict) -> None:
	"""Tiêu đề HAI dòng: dòng trên là ngày (cột ngày) hoặc nhãn cột, dòng dưới là thứ trong tuần.

	Trên màn hình datatable chỉ có một dòng tiêu đề nên ngày và thứ phải dồn chung một nhãn
	("1 T2"). Excel thì tách được, và tách ra mới đúng lối bảng chấm công VN: hàng số ngày, ngay
	dưới là hàng thứ. Cột không phải cột ngày gộp dọc qua cả hai dòng."""
	fill = PatternFill("solid", start_color="FF" + HEADER_BG)
	font = Font(bold=True, color="FF" + HEADER_FG)
	year, month = cint(filters.get("year")), cint(filters.get("month"))

	for i, col in enumerate(columns, start=1):
		fieldname = col.get("fieldname") or ""
		day = day_number(fieldname)
		top = ws.cell(HEADER_ROW, i, day if day else (col.get("label") or fieldname))
		bottom = ws.cell(WEEKDAY_ROW, i, weekday_label(year, month, day) if day else None)

		for cell in (top, bottom):
			cell.fill, cell.font, cell.alignment, cell.border = fill, font, HEADER_ALIGN, BORDER
		if not day:  # nhãn cột thường: gộp dọc để chữ nằm giữa hai dòng tiêu đề
			ws.merge_cells(start_row=HEADER_ROW, start_column=i, end_row=WEEKDAY_ROW, end_column=i)

	ws.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT
	ws.row_dimensions[WEEKDAY_ROW].height = 18


def day_number(fieldname: str) -> int | None:
	"""Số ngày trong tháng của một cột `day_<N>`; None nếu không phải cột ngày."""
	return cint(fieldname[4:]) if fieldname.startswith("day_") else None


def write_rows(ws, columns: list[dict], rows: list[dict], start_row: int) -> int:
	"""Ghi dữ liệu, trả về dòng cuối đã ghi (hoặc dòng ngay trước đó nếu bảng rỗng)."""
	for offset, data_row in enumerate(rows):
		write_row(ws, columns, data_row, start_row + offset, stt=offset + 1)
	return start_row + len(rows) - 1


def write_row(ws, columns: list[dict], data_row: dict, row: int, stt: int) -> None:
	"""Một dòng nhân viên: ô ngày tô theo `_state_<day>`, cột Tổng công in đậm."""
	for i, col in enumerate(columns, start=1):
		fieldname = col.get("fieldname")
		cell = ws.cell(row, i, stt if fieldname == "stt" else cell_value(data_row, col))
		cell.border = BORDER

		if fieldname.startswith("day_"):
			cell.alignment = CENTER
			fill, font = fill_for(data_row.get(f"_state_{fieldname[4:]}"))
			if fill:
				cell.fill, cell.font = fill, font
		elif col.get("fieldtype") in ("Float", "Int") or col.get("align") == "center":
			# `align` để cột chữ nhưng đọc như số (TB giờ/ngày dạng hh:mm) không bị dạt trái
			cell.alignment = CENTER
			cell.font = NUMBER_FONTS.get(fieldname, PLAIN_FONT)
		else:
			cell.alignment = LEFT


def cell_value(row: dict, col: dict):
	value = row.get(col.get("fieldname"))
	fieldtype = col.get("fieldtype")
	if fieldtype == "Float":
		return flt(value)
	if fieldtype == "Int":
		return cint(value)
	return value or ""


def set_widths(ws, columns: list[dict]) -> None:
	"""Cột ngày hẹp cố định; cột còn lại theo lưới, có sàn để nhãn xuống dòng được.

	Không suy bề rộng cột ngày từ `width` của report: trên màn hình cột đó phải nới ra cho vừa
	nhãn gộp "1 T2", còn ở đây thứ đã nằm ở dòng tiêu đề riêng nên ô chỉ cần chứa mã công."""
	for i, col in enumerate(columns, start=1):
		if day_number(col.get("fieldname") or ""):
			width = DAY_WIDTH
		else:
			# sàn chỉ nới tới mức nhãn cần: "STT" không việc gì phải rộng bằng "Ốm / chăm con ốm"
			label = str(col.get("label") or "")
			width = max(excel_width(col.get("width")), min(len(label) + 2, MIN_TEXT_WIDTH))
		ws.column_dimensions[get_column_letter(i)].width = width


def legend_span(ws, total_cols: int, groups: int) -> int:
	"""Số cột ngày gộp lại cho ô nghĩa: đủ rộng để đọc, nhưng vẫn để mọi cụm nằm trong bảng."""
	day_width = ws.column_dimensions[get_column_letter(FIRST_DAY_COLUMN)].width or 5.0
	desired = max(3, ceil(LEGEND_MEANING_CHARS / day_width))
	room = (total_cols - FIRST_DAY_COLUMN + 1) // max(groups, 1) - 1
	return max(3, min(desired, room))


def write_legend(ws, columns: list[dict], top: int) -> int:
	"""Khối chú thích dạng lưới: `Chú thích` một ô, mỗi ký hiệu một ô (tô đúng màu của nó), nghĩa
	ở ô ngang hàng ngay bên phải. Tối đa `MAX_LEGEND_ROWS` dòng — quá thì sang cụm cột kế tiếp.

	Trả về dòng cuối đã dùng (hoặc `top - 1` nếu không có ký hiệu nào) để khối trình ký biết
	đặt xuống đâu."""
	pairs = legend_pairs()
	if not pairs:
		return top - 1

	groups, rows = legend_layout(len(pairs))
	span = legend_span(ws, len(columns), groups)
	code_map = get_code_map()

	label = ws.cell(top, LEGEND_LABEL_COLUMN, LEGEND_LABEL)  # đúng MỘT ô, không gộp
	label.font = Font(bold=True)
	label.alignment = LEFT

	for index, (code, name) in enumerate(pairs):
		group, offset = divmod(index, rows)  # điền theo cột: đầy cụm này mới sang cụm kế
		row = top + offset
		col = FIRST_DAY_COLUMN + group * (span + 1)

		symbol = ws.cell(row, col, code)
		symbol.alignment, symbol.border = CENTER, BORDER
		fill, font = fill_for(day_state(code, code_map))
		symbol.font = font or Font(bold=True)
		if fill:
			symbol.fill = fill

		meaning = ws.cell(row, col + 1, name)
		meaning.alignment = LEFT
		ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + span)

	return top + rows - 1


def setup_print(ws, last_col: int) -> None:
	"""In ngang, co vừa một trang ngang, lặp cả hai dòng tiêu đề bảng ở mọi trang."""
	ws.page_setup.orientation = "landscape"
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0
	ws.sheet_properties.pageSetUpPr.fitToPage = True
	ws.print_title_rows = f"{HEADER_ROW}:{WEEKDAY_ROW}"
	ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"


@frappe.whitelist()
def download(filters=None, visible_idx=None, prepared_by=None, approved_by=None):
	"""Endpoint của nút Export (Excel có màu) trên `Monthly Attendance Report`."""
	from frappe.desk.utils import provide_binary_file

	# đúng thứ `export_query` kiểm — `ref_doctype` của báo cáo là Attendance
	frappe.permissions.can_export("Attendance", raise_exception=True)

	filters = frappe.parse_json(filters) or {}
	# Endpoint này chỉ phục vụ Bảng chấm công tháng. Nói thẳng ra khi bị gọi nhầm: `execute()` sẽ
	# ném "Please select month and year", đọc một mình không lần ra được là nút Export của báo cáo
	# NÀO gọi sai (đã mất công lần một lần, 2026-08-03).
	if not (cint(filters.get("month")) and cint(filters.get("year"))):
		frappe.throw(_("Thiếu tháng/năm: đường xuất Excel này chỉ dùng cho báo cáo Bảng chấm công tháng."))
	columns, data, _message = execute(filters)

	# lọc theo dòng đang hiện trên màn hình TRƯỚC khi bỏ dòng chú thích cũ: chỉ số client gửi lên
	# đánh trên `data` gốc, bỏ dòng trước thì lệch hết.
	visible = frappe.parse_json(visible_idx) if visible_idx else None
	if visible:
		keep = set(visible)
		data = [row for i, row in enumerate(data) if i in keep]

	stream = BytesIO()
	signatures = {"prepared_by": prepared_by, "approved_by": approved_by}
	build_workbook(columns, data, filters, signatures).save(stream)

	filters = frappe._dict(filters)
	name = f"Bang cham cong {cint(filters.month):02d}-{cint(filters.year)}"
	provide_binary_file(name, "xlsx", stream.getvalue())
