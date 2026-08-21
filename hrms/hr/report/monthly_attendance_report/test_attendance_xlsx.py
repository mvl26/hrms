# Copyright (c) 2026, Miyano Việt Nam.
"""File Excel xuất ra phải giống bảng trên màn hình: có màu, và chú thích là lưới chứ không phải
một dòng văn bản dài.

Test dựng workbook trong bộ nhớ (`build_workbook`) rồi soi từng ô — không đụng HTTP, không ghi
file, nên chạy được trong harness rollback."""

from io import BytesIO

import frappe
from frappe.tests.utils import FrappeTestCase
from frappe.utils import getdate

from hrms.hr.attendance_legend import legend_pairs
from hrms.hr.attendance_xlsx import (
	FIRST_DATA_ROW,
	FIRST_DAY_COLUMN,
	HEADER_ROW,
	LEGEND_LABEL,
	LEGEND_LABEL_COLUMN,
	MAX_LEGEND_ROWS,
	MIYANO_LETTERHEAD,
	SIGN_APPROVED_LABEL,
	SIGN_NAME_GAP,
	SIGN_PREPARED_LABEL,
	WEEKDAY_ROW,
	build_workbook,
	legend_layout,
	sign_blocks,
	signature_date_line,
)
from hrms.hr.report.monthly_attendance_report.monthly_attendance_report import STATE_STYLE, execute
from hrms.tests.isolation import PerTestRollback
from hrms.tests.vn_test_utils import test_employee


def rgb(cell) -> str:
	"""Mã màu nền của ô, viết thường không có alpha — `FFD9EFDC` → `d9efdc`."""
	fill = cell.fill
	if not fill or fill.fill_type != "solid":
		return ""
	return str(fill.start_color.rgb or "")[-6:].lower()


class TestAttendanceXlsx(PerTestRollback, FrappeTestCase):
	@classmethod
	def setUpClass(cls):
		super().setUpClass()
		cls.emp = test_employee()
		cls.year, cls.month = 2099, 3  # xa tương lai để không đụng dữ liệu thật

	def mk(self, day, **codes):
		att = frappe.get_doc(
			{
				"doctype": "Attendance",
				"employee": self.emp,
				"attendance_date": getdate(f"{self.year}-{self.month:02d}-{day:02d}"),
				**codes,
			}
		)
		att.insert()
		att.submit()
		return att

	def sheet(self, **extra):
		filters = {"month": self.month, "year": self.year, **extra}
		columns, data, _msg = execute(filters)
		wb = build_workbook(columns, data, filters)
		return wb.active

	def find_row(self, ws, employee):
		"""Dòng của nhân viên — dò theo cột tên, vì cột đầu giờ là STT chứ không phải mã NV."""
		name = frappe.db.get_value("Employee", employee, "employee_name")
		for row in ws.iter_rows(min_col=2, max_col=2, min_row=FIRST_DATA_ROW):
			if row[0].value == name:
				return row[0].row
		raise AssertionError(f"không thấy dòng của {employee} ({name}) trong sheet")

	def col_of(self, ws, header):
		"""Cột mang nhãn `header` trên dòng tiêu đề."""
		for row in ws.iter_rows():
			for cell in row:
				if cell.value == header:
					return cell.column
		raise AssertionError(f"không thấy cột {header}")

	def day_col(self, ws, day):
		"""Cột của ngày `day` — dòng tiêu đề trên mang SỐ ngày, dòng dưới mang thứ."""
		return self.col_of(ws, 1) - 1 + day

	# ── màu ô mã công ───────────────────────────────────────────────────────────────────────

	def test_day_cells_carry_state_colours(self):
		self.mk(5, custom_attendance_code="X")  # đi làm đủ  → work
		self.mk(6, custom_attendance_code="P")  # phép năm   → half (tím)
		self.mk(7, custom_attendance_code="V")  # vắng       → absent

		ws = self.sheet()
		r = self.find_row(ws, self.emp)

		self.assertEqual(rgb(ws.cell(r, self.day_col(ws, 5))), STATE_STYLE["work"]["bg"].lstrip("#"))
		self.assertEqual(rgb(ws.cell(r, self.day_col(ws, 6))), STATE_STYLE["half"]["bg"].lstrip("#"))
		self.assertEqual(rgb(ws.cell(r, self.day_col(ws, 7))), STATE_STYLE["absent"]["bg"].lstrip("#"))
		self.assertEqual(ws.cell(r, self.day_col(ws, 5)).value, "X")

	def test_half_day_cell_uses_half_colour(self):
		self.mk(11, custom_morning_code="X", custom_afternoon_code="P")

		ws = self.sheet()
		cell = ws.cell(self.find_row(ws, self.emp), self.day_col(ws, 11))
		self.assertEqual(cell.value, "X/P")
		self.assertEqual(rgb(cell), STATE_STYLE["half"]["bg"].lstrip("#"))

	def test_tong_cong_is_bold_and_matches_report(self):
		self.mk(5, custom_attendance_code="X")
		self.mk(6, custom_attendance_code="P")

		_cols, data, _msg = execute({"month": self.month, "year": self.year})
		expected = next(row["tong_cong"] for row in data if row.get("employee") == self.emp)

		ws = self.sheet()
		cell = ws.cell(self.find_row(ws, self.emp), self.col_of(ws, "Tổng công"))
		self.assertEqual(cell.value, expected)
		self.assertTrue(cell.font.bold, "cột Tổng công phải in đậm như trên màn hình")

	def test_header_is_frozen(self):
		ws = self.sheet()
		self.assertEqual(ws.freeze_panes, f"C{FIRST_DATA_ROW}")  # dưới HAI dòng tiêu đề bảng

	# ── tiêu đề thư ─────────────────────────────────────────────────────────────────────────

	def test_letterhead_block_is_left_aligned(self):
		ws = self.sheet(company=frappe.db.get_value("Company", {}, "name"))
		self.assertEqual(ws.cell(1, 1).value, MIYANO_LETTERHEAD["name"])
		self.assertTrue(ws.cell(1, 1).font.bold, "tên pháp nhân phải in đậm")
		self.assertIn(MIYANO_LETTERHEAD["tax_id"], str(ws.cell(2, 1).value))
		self.assertIn("Địa chỉ:", str(ws.cell(3, 1).value))
		for row in (1, 2, 3):
			self.assertEqual(ws.cell(row, 1).alignment.horizontal, "left")

	def test_title_and_period_are_centred_above_the_grid(self):
		ws = self.sheet()
		title_row, period_row = HEADER_ROW - 3, HEADER_ROW - 2
		self.assertEqual(ws.cell(title_row, 1).value, "BẢNG CHẤM CÔNG")
		self.assertEqual(ws.cell(period_row, 1).value, f"Tháng {self.month:02d} Năm {self.year}")
		for row in (title_row, period_row):
			self.assertEqual(ws.cell(row, 1).alignment.horizontal, "center")
		self.assertIsNone(ws.cell(HEADER_ROW - 1, 1).value, "phải có dòng trống trước bảng")

	def test_company_master_data_wins_over_the_fallback(self):
		"""Điền `tax_id` cho Company là số đó lên bản in ngay, không phải sửa code."""
		company = frappe.db.get_value("Company", {}, "name")
		frappe.db.set_value("Company", company, "tax_id", "9999999999")
		ws = self.sheet(company=company)
		self.assertIn("9999999999", str(ws.cell(2, 1).value))

	# ── cột STT ─────────────────────────────────────────────────────────────────────────────

	def test_first_column_is_a_running_number_not_the_employee_id(self):
		ws = self.sheet()
		self.assertEqual(ws.cell(HEADER_ROW, 1).value, "STT")

		numbers, row = [], FIRST_DATA_ROW
		while isinstance(ws.cell(row, 1).value, int):
			numbers.append(ws.cell(row, 1).value)
			row += 1
		self.assertTrue(numbers, "phải có ít nhất một dòng nhân viên")
		self.assertEqual(numbers, list(range(1, len(numbers) + 1)), "STT phải chạy 1, 2, 3…")

		self.assertFalse(
			[c for r in ws.iter_rows() for c in r if str(c.value or "").startswith("HR-EMP")],
			"mã nhân viên không được còn trong file",
		)

	def test_avg_office_hours_reaches_the_file(self):
		"""TB giờ/ngày phải theo được vào file, không chỉ có trên màn hình."""
		date = f"{self.year}-{self.month:02d}-04"
		att = frappe.get_doc(
			{
				"doctype": "Attendance",
				"employee": self.emp,
				"attendance_date": getdate(date),
				"custom_attendance_code": "X",
				"in_time": f"{date} 08:00:00",
				"out_time": f"{date} 17:30:00",
			}
		)
		att.insert()
		att.submit()

		ws = self.sheet()
		cell = ws.cell(self.find_row(ws, self.emp), self.col_of(ws, "TB giờ/ngày"))
		self.assertEqual(cell.value, "08:00")  # 9.5h trừ 1.5h nghỉ trưa, đọc theo đồng hồ
		self.assertEqual(cell.alignment.horizontal, "center", "cột đọc như số thì không dạt trái")

	# ── thứ trong tuần ──────────────────────────────────────────────────────────────────────

	def test_weekday_row_sits_under_the_day_numbers(self):
		"""Excel tách được nên xếp đúng lối bảng chấm công VN: hàng số ngày, ngay dưới là hàng thứ."""
		filters = {"month": 3, "year": 2026}  # 2026-03-01 là Chủ nhật
		columns, data, _msg = execute(filters)
		ws = build_workbook(columns, data, filters).active

		first_day = self.col_of(ws, 1)
		self.assertEqual(ws.cell(HEADER_ROW, first_day).value, 1)
		self.assertEqual(ws.cell(WEEKDAY_ROW, first_day).value, "CN")
		self.assertEqual(ws.cell(WEEKDAY_ROW, first_day + 1).value, "T2")
		self.assertEqual(ws.cell(WEEKDAY_ROW, first_day + 6).value, "T7")  # ngày 7
		self.assertEqual(ws.cell(WEEKDAY_ROW, first_day + 7).value, "CN")  # ngày 8, tròn một tuần

	def test_plain_columns_span_both_header_rows(self):
		"""Cột không phải cột ngày gộp dọc, không để nhãn lửng lơ trên hàng thứ."""
		ws = self.sheet()
		merged = {str(rng) for rng in ws.merged_cells.ranges}
		self.assertIn(f"A{HEADER_ROW}:A{WEEKDAY_ROW}", merged, "cột STT phải gộp qua hai dòng tiêu đề")
		self.assertIn(f"B{HEADER_ROW}:B{WEEKDAY_ROW}", merged, "cột Nhân viên phải gộp qua hai dòng")
		self.assertIsNone(ws.cell(WEEKDAY_ROW, self.col_of(ws, "Tổng công")).value)

	# ── khối chú thích ──────────────────────────────────────────────────────────────────────

	def test_no_long_text_legend_anywhere(self):
		"""Không ô nào được dồn cả danh sách mã kiểu `X=Đi làm đủ công; ...` — đã có khối lưới."""
		ws = self.sheet()
		for row in ws.iter_rows():
			for cell in row:
				if isinstance(cell.value, str):
					self.assertNotIn("X=", cell.value, "dòng chú thích dạng văn bản dài còn sót")

	def test_legend_grid_layout(self):
		ws = self.sheet()
		pairs = legend_pairs()

		label_cells = [c for row in ws.iter_rows() for c in row if c.value == LEGEND_LABEL]
		self.assertEqual(len(label_cells), 1, "chữ 'Chú thích' phải nằm đúng MỘT ô")
		self.assertEqual(label_cells[0].column, LEGEND_LABEL_COLUMN, "phải nằm ở cột Nhân viên")
		top = label_cells[0].row

		# mỗi ký hiệu một ô, nghĩa ở ô ngang hàng ngay bên phải
		seen = {}
		for row in ws.iter_rows(min_row=top):
			for cell in row:
				if cell.value in dict(pairs) and cell.column >= FIRST_DAY_COLUMN:
					seen[cell.value] = (cell.row, cell.column)
		self.assertEqual(set(seen), {code for code, _name in pairs}, "thiếu ký hiệu trong chú thích")

		for code, name in pairs:
			r, c = seen[code]
			self.assertEqual(ws.cell(r, c + 1).value, name, f"nghĩa của {code} phải ngang hàng bên phải")

		rows_used = {r for r, _c in seen.values()}
		self.assertLessEqual(len(rows_used), MAX_LEGEND_ROWS, "chú thích không được quá 10 dòng")

	def test_legend_symbol_cells_are_coloured(self):
		ws = self.sheet()
		codes = {}
		for row in ws.iter_rows():
			for cell in row:
				if cell.value in ("X", "P", "V") and cell.column > 2 and rgb(cell):
					codes.setdefault(cell.value, []).append(rgb(cell))
		# ký hiệu trong chú thích mang đúng màu state của nó (giống hệt ô trong lưới)
		self.assertIn(STATE_STYLE["work"]["bg"].lstrip("#"), codes.get("X", []))
		self.assertIn(STATE_STYLE["half"]["bg"].lstrip("#"), codes.get("P", []))
		self.assertIn(STATE_STYLE["absent"]["bg"].lstrip("#"), codes.get("V", []))

	# ── endpoint của nút Export ─────────────────────────────────────────────────────────────

	def test_download_returns_a_readable_xlsx(self):
		"""Đi hết đường thật: whitelisted method → `frappe.response` mang file mở lại được."""
		from openpyxl import load_workbook

		from hrms.hr.attendance_xlsx import download

		self.mk(5, custom_attendance_code="X")
		response = frappe.local.response
		try:
			download(filters=frappe.as_json({"month": self.month, "year": self.year}))
			self.assertEqual(frappe.response["type"], "binary")
			self.assertTrue(frappe.response["filename"].endswith(".xlsx"))
			ws = load_workbook(BytesIO(frappe.response["filecontent"])).active
			period = str(ws.cell(HEADER_ROW - 2, 1).value)
			self.assertEqual(period, f"Tháng {self.month:02d} Năm {self.year}")
		finally:
			frappe.local.response = response

	def test_download_says_plainly_when_called_without_a_period(self):
		"""Bị gọi nhầm từ báo cáo khác (Salary Register, 2026-08-03) thì phải nói rõ là gọi nhầm.

		`execute()` ném "Please select month and year" — đọc log một mình không lần ra được là nút
		Export của báo cáo NÀO gọi sai."""
		from hrms.hr.attendance_xlsx import download

		with self.assertRaises(frappe.ValidationError) as caught:
			download(filters=frappe.as_json({}))
		self.assertIn("Bảng chấm công tháng", str(caught.exception))

	# ── khối ký tên cuối bảng ───────────────────────────────────────────────────────────────

	def sign_rows(self, ws):
		"""(dòng ngày tháng, dòng chức danh, dòng tên) của khối ký cuối bảng."""
		for row in ws.iter_rows(min_row=FIRST_DATA_ROW):
			for cell in row:
				if cell.value == SIGN_PREPARED_LABEL:
					return cell.row - 1, cell.row, cell.row + SIGN_NAME_GAP
		raise AssertionError("không thấy khối ký tên cuối bảng")

	def merged_at(self, ws, row, col) -> str:
		"""Vùng gộp chứa ô (row, col) — chuỗi rỗng nếu ô không nằm trong vùng nào."""
		for rng in ws.merged_cells.ranges:
			if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
				return str(rng)
		return ""

	def test_signature_block_sits_below_the_legend(self):
		"""Đủ ba dòng như biểu mẫu gốc: địa danh + ngày, hai chức danh, rồi chỗ ký."""
		ws = self.sheet()
		legend_row = next(c.row for row in ws.iter_rows() for c in row if c.value == LEGEND_LABEL)
		date_row, title_row, name_row = self.sign_rows(ws)
		self.assertGreater(date_row, legend_row, "khối ký phải nằm dưới khối chú thích")

		last_col = max(c.column for c in ws[HEADER_ROW] if c.value is not None)
		prepared, approved = sign_blocks(last_col)

		title = ws.cell(title_row, prepared[0])
		self.assertEqual(title.value, SIGN_PREPARED_LABEL)
		self.assertTrue(title.font.bold, "chức danh phải in đậm")
		self.assertEqual(title.alignment.horizontal, "center")

		approver = ws.cell(title_row, approved[0])
		self.assertEqual(approver.value, SIGN_APPROVED_LABEL)
		self.assertEqual(approved[1], last_col, "khối Người duyệt phải sát mép phải bảng")
		self.assertEqual(ws.cell(date_row, approved[0]).value, signature_date_line(None))
		self.assertTrue(ws.cell(date_row, approved[0]).font.italic, "dòng ngày tháng in nghiêng")

		# mỗi chức danh gộp cả khối cột của nó → chữ nằm giữa chỗ ký
		self.assertTrue(self.merged_at(ws, title_row, prepared[0]), "Người lập phải gộp ô")
		self.assertTrue(self.merged_at(ws, title_row, approved[0]), "Người duyệt phải gộp ô")

		self.assertGreater(name_row, title_row, "phải chừa chỗ ký giữa chức danh và tên")
		for row in range(title_row + 1, name_row):
			self.assertFalse(
				[c.value for c in ws[row] if c.value not in (None, "")],
				f"dòng {row} phải để trống làm chỗ ký",
			)

	def test_signature_date_is_the_export_day_in_the_company_city(self):
		"""`Hà Nội, ngày 02 tháng 7 năm 2026` — địa danh của công ty, ngày xuất file."""
		from frappe.utils import getdate, nowdate

		today = getdate(nowdate())
		self.assertEqual(
			signature_date_line(None),
			f"{MIYANO_LETTERHEAD['city']}, ngày {today.day:02d} tháng {today.month} năm {today.year}",
		)

	def test_signature_names_land_under_their_titles(self):
		filters = {"month": self.month, "year": self.year}
		columns, data, _msg = execute(filters)
		ws = build_workbook(
			columns,
			data,
			filters,
			signatures={"prepared_by": "Phan Thị Thu Lan", "approved_by": "Đoàn Ngọc Anh"},
		).active

		_date_row, title_row, name_row = self.sign_rows(ws)
		last_col = max(c.column for c in ws[HEADER_ROW] if c.value is not None)
		prepared, approved = sign_blocks(last_col)

		self.assertEqual(ws.cell(name_row, prepared[0]).value, "Phan Thị Thu Lan")
		self.assertEqual(ws.cell(name_row, approved[0]).value, "Đoàn Ngọc Anh")
		self.assertEqual(ws.cell(name_row, prepared[0]).alignment.horizontal, "center")
		self.assertEqual(name_row, title_row + SIGN_NAME_GAP)

	def test_signature_names_are_blank_when_nobody_is_named(self):
		"""Không ai được điền thì để trống cho ký tay — không in `Administrator` lên bản in."""
		ws = self.sheet()
		_date_row, _title_row, name_row = self.sign_rows(ws)
		self.assertFalse([c.value for c in ws[name_row] if c.value not in (None, "")])

	def test_sign_blocks_stay_inside_the_table(self):
		"""Bảng rộng hẹp thế nào hai khối cũng nằm trong bảng và không chồng lên nhau."""
		for last_col in range(FIRST_DAY_COLUMN + 2, 60):
			prepared, approved = sign_blocks(last_col)
			self.assertEqual(approved[1], last_col)
			self.assertGreaterEqual(prepared[0], FIRST_DAY_COLUMN, f"tràn sang cột tên ({last_col})")
			self.assertLess(prepared[1], approved[0], f"hai khối chồng nhau ({last_col})")

	def test_download_carries_the_signature_names(self):
		from openpyxl import load_workbook

		from hrms.hr.attendance_xlsx import download

		response = frappe.local.response
		try:
			download(
				filters=frappe.as_json({"month": self.month, "year": self.year}),
				prepared_by="Phan Thị Thu Lan",
				approved_by="Đoàn Ngọc Anh",
			)
			ws = load_workbook(BytesIO(frappe.response["filecontent"])).active
		finally:
			frappe.local.response = response

		names = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
		self.assertIn("Phan Thị Thu Lan", names)
		self.assertIn("Đoàn Ngọc Anh", names)

	# ── luật chia cụm cột (thuần hàm, không cần dữ liệu) ─────────────────────────────────────

	def test_legend_splits_into_column_groups(self):
		self.assertEqual(legend_layout(8), (1, 8))  # vừa 10 dòng → một cụm
		self.assertEqual(legend_layout(10), (1, 10))  # đúng ngưỡng → vẫn một cụm
		self.assertEqual(legend_layout(16), (2, 8))  # 16 mã → 2 cụm, 8 dòng
		self.assertEqual(legend_layout(21), (3, 7))  # 21 mã → 3 cụm, 7 dòng
		for n in range(1, 60):
			groups, rows = legend_layout(n)
			self.assertLessEqual(rows, MAX_LEGEND_ROWS)
			self.assertGreaterEqual(groups * rows, n, "phải đủ chỗ cho mọi ký hiệu")
