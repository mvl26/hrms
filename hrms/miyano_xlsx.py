# Copyright (c) 2026, Miyano Việt Nam.
"""Kiểu nhà Miyano cho mọi biểu mẫu Excel xuất từ hệ thống — tiêu đề thư và khối trình ký.

Bảng chấm công (`hrms/hr/attendance_xlsx.py`) và bảng lương (`hrms/vn_payroll/salary_xlsx.py`) là
hai báo cáo khác hẳn nhau về cột, nhưng **đầu và cuối tờ giấy thì giống hệt**: ba dòng pháp nhân
căn trái, tên biểu mẫu + kỳ căn giữa, rồi cuối bảng là dòng địa danh/ngày tháng và hai khối ký.
Phần đó nằm ở đây để hai bên không trôi khỏi nhau — sửa địa chỉ công ty một chỗ, cả hai biểu mẫu
cùng đổi.

Chỉ dựng ô, không đọc dữ liệu nghiệp vụ nào ngoài Company/Address → không có tác dụng phụ.
"""

from openpyxl.styles import Alignment, Border, Font, Side

import frappe
from frappe import _
from frappe.utils import cint, getdate, nowdate

# Tiêu đề thư của Miyano. Deployment này chỉ có MỘT pháp nhân (xem CLAUDE.md), và không dòng nào
# trong ba dòng dưới đây có sẵn trong master data: `Company.company_name` là tên gọi tắt "Miyano",
# `tax_id` để trống, site chưa có Address nào. Để đây thay vì sửa `company_name` vì đổi tên đó là
# đổi tên công ty trên MỌI chứng từ ERPNext, không chỉ một biểu mẫu. `company_lines()` vẫn ưu tiên
# master data khi có, nên điền `tax_id` / tạo Address là hết dùng tới mặc định này.
MIYANO_LETTERHEAD = {
	"name": "CÔNG TY TNHH MIYANO VIỆT NAM",
	"tax_id": "0109529507",
	"address": ("số 20, Khu C17, ngõ 264/63, đường Ngọc Thụy, Phường Bồ Đề, Thành phố Hà Nội, Việt Nam"),
	# địa danh mở đầu dòng ngày tháng của khối trình ký ("Hà Nội, ngày 02 tháng 7 năm 2026")
	"city": "Hà Nội",
}

HEADER_BG = "E9EDF2"
HEADER_FG = "1F272E"
GRID_LINE = "D1D8DD"

THIN = Side(style="thin", color=GRID_LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
# Nhãn cột thường dài hơn ô chứa nó. Xuống dòng trong ô thay vì nới cột: nới ra thì cả bảng bị co
# nhỏ khi in vừa một trang ngang.
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_HEIGHT = 34  # đủ hai dòng chữ ở cỡ mặc định

# Khối trình ký cuối bảng — bản Excel gốc của Miyano (`docs/2. Bang_Cham_Cong_06-2026_final.xlsx`)
# xếp cả hai chữ ký ở NỬA PHẢI: người duyệt sát mép phải, người lập lùi vào giữa.
SIGN_PREPARED_LABEL = "Người lập"
SIGN_APPROVED_LABEL = "Người duyệt"

# Bề ngang mỗi khối ký, tính bằng cột (bản gốc gộp 10 cột ngày: AF..AO). Bảng hẹp thì co lại.
SIGN_BLOCK_WIDTH = 10

# Số dòng từ chức danh xuống dòng tên — chỗ trống để ký tay (bản gốc: dòng 28 → dòng 34).
SIGN_NAME_GAP = 6


def excel_width(px, minimum: float = 4.0) -> float:
	"""Đổi bề rộng cột của datatable (px) sang đơn vị ký tự của Excel."""
	return max(minimum, round(cint(px or 100) / 8.0, 1))


# ---------------------------------------------------------------- tiêu đề thư


def company_lines(company: str | None) -> list[str]:
	"""Ba dòng tiêu đề thư: tên pháp nhân, MST, địa chỉ. Bỏ qua dòng nào không có dữ liệu.

	Master data thắng ở đâu có: `Company.tax_id` cho MST, Address chính liên kết với Company cho
	địa chỉ. Site hiện chưa có cả hai (2026-08-03: `tax_id` trống, không Address nào) và
	`Company.company_name` là tên gọi tắt "Miyano" chứ không phải tên pháp nhân trên giấy tờ, nên
	`MIYANO_LETTERHEAD` đứng làm mặc định — điền vào Company/Address là dữ liệu thắng ngay, không
	phải sửa code."""
	tax_id = frappe.db.get_value("Company", company, "tax_id") if company else None
	lines = [MIYANO_LETTERHEAD["name"]]
	if tax_id or MIYANO_LETTERHEAD["tax_id"]:
		lines.append(_("MST: {0}").format(tax_id or MIYANO_LETTERHEAD["tax_id"]))
	address = company_address(company) or MIYANO_LETTERHEAD["address"]
	if address:
		lines.append(_("Địa chỉ: {0}").format(address))
	return lines


def company_address_row(company: str | None) -> frappe._dict | None:
	"""Address chính liên kết với Company; None nếu chưa có Address nào."""
	if not company:
		return None
	names = frappe.get_all(
		"Dynamic Link",
		filters={"link_doctype": "Company", "link_name": company, "parenttype": "Address"},
		pluck="parent",
	)
	if not names:
		return None
	addresses = frappe.get_all(
		"Address",
		filters={"name": ["in", names]},
		fields=["address_line1", "address_line2", "city", "state", "country", "is_primary_address"],
		order_by="is_primary_address desc",
		limit=1,
	)
	return addresses[0] if addresses else None


def company_address(company: str | None) -> str | None:
	"""Địa chỉ chính của Company gộp thành một dòng; None nếu chưa có Address nào liên kết."""
	a = company_address_row(company)
	if not a:
		return None
	return ", ".join(p for p in (a.address_line1, a.address_line2, a.city, a.state, a.country) if p)


def company_city(company: str | None) -> str | None:
	"""Địa danh đứng đầu dòng ngày tháng — lấy `city` của Address chính khi có."""
	a = company_address_row(company)
	return (a.city or None) if a else None


def period_line(filters: dict) -> str:
	"""Dòng kỳ dưới tên biểu mẫu: `Tháng 06 Năm 2026`."""
	return _("Tháng {0} Năm {1}").format(f"{cint(filters.get('month')):02d}", cint(filters.get("year")))


def write_letterhead(ws, company: str | None, last_col: int, title: str, subtitle: str, title_row: int):
	"""Khối pháp nhân căn TRÁI ở đầu sheet, rồi tên biểu mẫu + kỳ căn GIỮA tại `title_row`.

	Mỗi dòng gộp hết bề ngang bảng để khi in không bị cắt ở mép cột. `title_row` neo cứng chứ
	không suy từ số dòng pháp nhân vừa ghi: thiếu MST hay địa chỉ thì khối trên ngắn lại, nhưng
	bảng vẫn phải bắt đầu đúng chỗ mọi nơi khác trông đợi."""
	for row, text in enumerate(company_lines(company), start=1):
		cell = ws.cell(row, 1, text)
		cell.font = Font(bold=(row == 1), size=11)
		cell.alignment = LEFT
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)

	for offset, (text, font) in enumerate(((title, Font(bold=True, size=16)), (subtitle, Font(size=12)))):
		cell = ws.cell(title_row + offset, 1, text)
		cell.font, cell.alignment = font, CENTER
		ws.merge_cells(
			start_row=title_row + offset, start_column=1, end_row=title_row + offset, end_column=last_col
		)
	ws.row_dimensions[title_row].height = 26


# ------------------------------------------------------------- khối trình ký


def sign_blocks(
	last_col: int, first_col: int = 3, width: int = SIGN_BLOCK_WIDTH
) -> tuple[tuple[int, int], tuple[int, int]]:
	"""(cột đầu, cột cuối) của khối "Người lập" và khối "Người duyệt".

	Cả hai nằm ở nửa phải bảng như biểu mẫu gốc: người duyệt sát mép phải, người lập lùi vào giữa,
	giữa hai khối chừa một quãng trống. Bảng ít cột thì khối co lại chứ không tràn sang cột tên.

	`first_col` là cột trái nhất khối ký được phép chạm tới (sau các cột định danh), `width` là bề
	ngang mong muốn — bảng chấm công cột ngày hẹp nên gộp 10 cột, bảng lương cột tiền rộng nên ít
	hơn nhiều."""
	room = max(last_col - first_col + 1, 2)
	# 2 khối + ít nhất 1 cột ngăn cách phải lọt trong `room`
	width = max(1, min(width, (room - 1) // 2))
	approved = (last_col - width + 1, last_col)
	start = max(first_col, approved[0] - width - max(1, width // 2))
	return (start, start + width - 1), approved


def signature_date_line(company: str | None) -> str:
	"""Dòng trên chữ ký người duyệt: `Hà Nội, ngày 02 tháng 7 năm 2026`.

	Ngày là ngày xuất file (ngày trình ký), địa danh lấy `city` của Address công ty khi có."""
	today = getdate(nowdate())
	city = company_city(company) or MIYANO_LETTERHEAD["city"]
	return _("{0}, ngày {1} tháng {2} năm {3}").format(city, f"{today.day:02d}", today.month, today.year)


def signature_names(signatures: dict | None) -> tuple[str, str]:
	"""Tên dưới hai chữ ký. Người lập bỏ trống thì lấy người đang xuất file; người duyệt thì không
	đoán — ai duyệt là việc của người trình ký, để trống cho ký tay."""
	signatures = frappe._dict(signatures or {})
	prepared = (signatures.prepared_by or "").strip() or session_signer_name()
	return prepared, (signatures.approved_by or "").strip()


def session_signer_name() -> str:
	"""Tên người đang đăng nhập, theo hồ sơ nhân viên rồi mới tới tên User.

	`Administrator` / `Guest` là tài khoản hệ thống — in mấy chữ đó lên bản trình ký thì vô nghĩa,
	thà để trống ô tên cho người ký điền tay."""
	user = frappe.session.user
	if not user or user in ("Guest", "Administrator"):
		return ""
	return (
		frappe.db.get_value("Employee", {"user_id": user, "status": "Active"}, "employee_name")
		or frappe.db.get_value("User", user, "full_name")
		or ""
	)


def write_signatures(
	ws,
	last_col: int,
	top: int,
	signatures: dict | None,
	company: str | None,
	first_col: int = 3,
	width: int = SIGN_BLOCK_WIDTH,
) -> None:
	"""Khối trình ký cuối bảng: dòng địa danh + ngày, hai chức danh, chừa chỗ ký rồi tới tên.

	Không kẻ khung: đây là chỗ ký tay trên bản in, kẻ ô vào chỉ vướng chữ ký."""
	prepared_block, approved_block = sign_blocks(last_col, first_col, width)
	prepared_name, approved_name = signature_names(signatures)

	date_cell = ws.cell(top, approved_block[0], signature_date_line(company))
	date_cell.font, date_cell.alignment = Font(italic=True), CENTER
	merge_across(ws, top, approved_block)

	name_row = top + 1 + SIGN_NAME_GAP
	for block, label, name in (
		(prepared_block, SIGN_PREPARED_LABEL, prepared_name),
		(approved_block, SIGN_APPROVED_LABEL, approved_name),
	):
		title = ws.cell(top + 1, block[0], label)
		title.font, title.alignment = Font(bold=True), CENTER
		merge_across(ws, top + 1, block)

		signer = ws.cell(name_row, block[0], name or None)
		signer.alignment = CENTER
		merge_across(ws, name_row, block)


def merge_across(ws, row: int, block: tuple[int, int]) -> None:
	"""Gộp một dòng qua cả khối cột — bỏ qua khối rộng đúng một cột (openpyxl không nhận)."""
	if block[1] > block[0]:
		ws.merge_cells(start_row=row, start_column=block[0], end_row=row, end_column=block[1])
