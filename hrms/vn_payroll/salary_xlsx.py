# Copyright (c) 2026, Miyano Việt Nam.
"""Xuất bảng lương tháng ra Excel theo biểu mẫu Miyano — có tiêu đề thư và khối trình ký.

Nút Export mặc định của query report đi qua `frappe.desk.query_report.export_query`: nó đổ thẳng
`columns` + `result` qua `make_xlsx()`, ra một lưới trần không tiêu đề công ty, không dòng tổng in
đậm, không chỗ ký. Bảng lương thì phải KÝ được — đó là chứng từ trình giám đốc, không phải bản kết
xuất dữ liệu. Module này là đường xuất riêng của báo cáo `MVL Salary Register`, dựng workbook bằng
openpyxl nên kiểm soát được từng ô.

Hai đường ra:

- `build_workbook(columns, data, filters, signatures)` — thuần hàm, không đụng HTTP, để test soi ô.
- `download(filters, visible_idx, prepared_by, approved_by)` — endpoint cho nút Export trên report.

Đầu và cuối tờ giấy dùng chung với bảng chấm công (`hrms/miyano_xlsx.py`) → sửa địa chỉ công ty
một chỗ, cả hai biểu mẫu cùng đổi.

CHỈ ĐỌC: không đụng Salary Slip, không tính lại đồng nào — mọi con số lấy nguyên từ `execute()` của
báo cáo, mà báo cáo lấy nguyên từ `Salary Detail` đã chốt trên phiếu. Lương bất biến theo định nghĩa.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import frappe
from frappe import _
from frappe.utils import cint

from hrms.miyano_xlsx import (
	BORDER,
	CENTER,
	HEADER_ALIGN,
	HEADER_BG,
	HEADER_FG,
	LEFT,
	RIGHT,
	excel_width,
	period_line,
	write_letterhead,
	write_signatures,
)
from hrms.payroll.report.mvl_salary_register.mvl_salary_register import TOTAL_LABEL, execute

SHEET_TITLE = "BẢNG THANH TOÁN TIỀN LƯƠNG"

# Tên tab của sheet (Excel giới hạn 31 ký tự) — ngắn hơn tiêu đề in trên giấy.
SHEET_NAME = "Bảng lương"

# Bảng lương chỉ có MỘT dòng tiêu đề (bảng chấm công cần hai: số ngày rồi thứ). Neo bằng hằng số
# chứ không đếm theo số dòng pháp nhân thực ghi: thiếu MST hay địa chỉ thì khối trên ngắn lại,
# bảng vẫn phải bắt đầu đúng chỗ khối ký và `print_title_rows` trông đợi.
HEADER_ROW = 8
FIRST_DATA_ROW = HEADER_ROW + 1

# Cột STT thay cột "Mã NV" của báo cáo: bản in cần số thứ tự, mã hệ thống (HR-EMP-00001) chỉ tốn
# chỗ. Không đổi `columns` của report vì trên màn hình cột Mã NV là liên kết bấm sang hồ sơ.
STT_COLUMN = {"fieldname": "stt", "label": "STT", "fieldtype": "Int", "width": 44}

# Tiền VND không có phần lẻ; dấu phân cách nghìn là thứ khiến bảng đọc được. `25994783` thì không
# ai soát nổi, `25,994,783` thì liếc qua là biết bậc số.
MONEY_FORMAT = "#,##0"
COEFFICIENT_FORMAT = "0.00"

# Dấu cảnh báo trên tiêu đề khi bảng có lẫn phiếu chưa chốt. Không có nó thì tờ giấy trình ký
# trông y hệt bản chính thức — người duyệt ký lên những con số còn có thể đổi mà không biết.
DRAFT_WARNING = "GỒM CẢ PHIẾU NHÁP — CHƯA CHỐT"

TOTAL_FILL = PatternFill("solid", start_color="FFF1F5F9")
TOTAL_FONT = Font(bold=True)
HEADER_HEIGHT = 46  # nhãn cột dài ("TN chịu thuế kê khai (U)") cần ba dòng ở bề rộng này

# Bề rộng tối thiểu của cột tiền: đủ chỗ cho `28,584,783` mà không phải nới tay.
MIN_MONEY_WIDTH = 15.0
MIN_TEXT_WIDTH = 11.0

# Khối ký gộp 4 cột — cột tiền ở đây rộng gấp ba cột ngày của bảng chấm công, gộp 10 cột như bên
# đó thì hai chữ ký chiếm gần hết bề ngang tờ giấy.
SIGN_WIDTH = 4
# Cột trái nhất khối ký được phép chạm tới: sau STT + Họ tên.
SIGN_FIRST_COLUMN = 3


def build_workbook(
	columns: list[dict],
	data: list[dict],
	filters: dict | None = None,
	signatures: dict | None = None,
) -> Workbook:
	"""Dựng workbook bảng lương từ đúng `columns`/`data` mà `execute()` của báo cáo trả về.

	`signatures` là hai cái tên trên khối trình ký (`prepared_by` / `approved_by`) — người xuất file
	điền ở hộp thoại Export; để trống thì chỉ in chức danh, ký tên bằng tay."""
	filters = frappe._dict(filters or {})
	columns = excel_columns(columns)
	body, total = split_total_row(data)

	wb = Workbook()
	ws = wb.active
	ws.title = _(SHEET_NAME)

	last_col = len(columns)
	write_letterhead(
		ws,
		filters.get("company"),
		last_col,
		_(SHEET_TITLE),
		subtitle_line(filters),
		title_row=HEADER_ROW - 3,  # tên bảng, kỳ lương, rồi một dòng trống trước tiêu đề bảng
	)
	write_header(ws, columns)
	end_row = write_rows(ws, columns, body, total, FIRST_DATA_ROW)
	set_widths(ws, columns)

	# đóng băng dưới tiêu đề, bên phải cột họ tên: cuộn sang cột thuế/BH vẫn biết đang xem ai
	ws.freeze_panes = f"{get_column_letter(SIGN_FIRST_COLUMN)}{FIRST_DATA_ROW}"

	write_signatures(
		ws, last_col, end_row + 2, signatures, filters.get("company"), SIGN_FIRST_COLUMN, SIGN_WIDTH
	)
	setup_print(ws, last_col)
	return wb


def subtitle_line(filters: dict) -> str:
	"""Dòng dưới tên biểu mẫu: kỳ lương, kèm cảnh báo nếu bảng có lẫn phiếu nháp."""
	line = period_line(filters)
	if cint(filters.get("include_drafts")):
		line = f"{line} — {_(DRAFT_WARNING)}"
	return line


def excel_columns(columns: list[dict]) -> list[dict]:
	"""Cột của file: thay "Mã NV" bằng STT, giữ nguyên phần còn lại của báo cáo."""
	return [STT_COLUMN if c.get("fieldname") == "employee" else c for c in columns]


def split_total_row(data: list[dict]) -> tuple[list[dict], dict | None]:
	"""Tách dòng TỔNG CỘNG khỏi các dòng nhân viên.

	Nhận diện bằng CHỖ THIẾU `employee`, không so chuỗi nhãn: nhãn đi qua `_()` nên đổi theo ngôn
	ngữ, còn dòng tổng thì vĩnh viễn không có mã nhân viên. Lọc kiểu `if row.get("employee")` như
	bảng chấm công sẽ ném luôn dòng tổng đi — đúng thứ người ta nhìn đầu tiên trên bảng lương."""
	body = [r for r in data if r and r.get("employee")]
	totals = [r for r in data if r and not r.get("employee") and r.get("employee_name")]
	return body, (totals[-1] if totals else None)


def write_header(ws, columns: list[dict]) -> None:
	"""Một dòng tiêu đề, chữ trắng nền xám, nhãn dài thì xuống dòng trong ô."""
	fill = PatternFill("solid", start_color="FF" + HEADER_BG)
	font = Font(bold=True, color="FF" + HEADER_FG)
	for i, col in enumerate(columns, start=1):
		cell = ws.cell(HEADER_ROW, i, col.get("label") or col.get("fieldname"))
		cell.fill, cell.font, cell.alignment, cell.border = fill, font, HEADER_ALIGN, BORDER
	ws.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT


def write_rows(ws, columns: list[dict], body: list[dict], total: dict | None, start_row: int) -> int:
	"""Ghi các dòng nhân viên rồi dòng TỔNG CỘNG; trả về dòng cuối đã ghi."""
	row = start_row - 1
	for offset, data_row in enumerate(body):
		row = start_row + offset
		write_row(ws, columns, data_row, row, stt=offset + 1)
	if total is not None:
		row += 1
		write_row(ws, columns, total, row, stt=None, bold=True)
	return row


def write_row(ws, columns: list[dict], data_row: dict, row: int, stt: int | None, bold: bool = False):
	"""Một dòng dữ liệu. `stt=None` là dòng tổng — không đánh số thứ tự cho nó."""
	for i, col in enumerate(columns, start=1):
		fieldname = col.get("fieldname") or ""
		value = stt if fieldname == "stt" else data_row.get(fieldname)
		cell = ws.cell(row, i, value if value != "" else None)
		cell.border = BORDER
		cell.number_format = number_format(col)
		cell.alignment = alignment_for(col)
		if bold:
			cell.font, cell.fill = TOTAL_FONT, TOTAL_FILL


def number_format(col: dict) -> str:
	"""Tiền có dấu phân cách nghìn; hệ số hai chữ số; còn lại để `General`.

	Số công để `General` là cố ý: 19.5 hiện `19.5`, 22 hiện `22` — ép `0.00` thì cả cột đầy
	`22.00`, đọc rối mà chẳng thêm thông tin gì."""
	if col.get("fieldtype") == "Currency":
		return MONEY_FORMAT
	if col.get("fieldname") == "coefficient":
		return COEFFICIENT_FORMAT
	return "General"


def alignment_for(col: dict):
	"""Tiền căn phải (so bậc số bằng mắt), STT/số căn giữa, chữ căn trái."""
	if col.get("fieldtype") == "Currency":
		return RIGHT
	if col.get("fieldtype") in ("Data", "Link"):
		return LEFT
	return CENTER


def set_widths(ws, columns: list[dict]) -> None:
	"""Bề rộng theo lưới trên màn hình, có sàn riêng cho cột tiền."""
	for i, col in enumerate(columns, start=1):
		floor = MIN_MONEY_WIDTH if col.get("fieldtype") == "Currency" else MIN_TEXT_WIDTH
		label = str(col.get("label") or "")
		# sàn chỉ nới tới mức nhãn cần: "STT" không việc gì phải rộng bằng "Thực lĩnh (T)"
		width = max(excel_width(col.get("width")), min(len(label) + 2, floor))
		ws.column_dimensions[get_column_letter(i)].width = width


def setup_print(ws, last_col: int) -> None:
	"""In ngang, co vừa một trang ngang, lặp dòng tiêu đề bảng ở mọi trang."""
	ws.page_setup.orientation = "landscape"
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0
	ws.sheet_properties.pageSetUpPr.fitToPage = True
	ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"
	ws.print_area = f"A1:{get_column_letter(last_col)}{ws.max_row}"


@frappe.whitelist()
def download(
	filters: str | dict | None = None,
	visible_idx: str | list | None = None,
	prepared_by: str | None = None,
	approved_by: str | None = None,
):
	"""Endpoint của nút Export (Excel có tiêu đề + khối ký) trên `MVL Salary Register`."""
	from frappe.desk.utils import provide_binary_file

	# đúng thứ `export_query` kiểm — `ref_doctype` của báo cáo là Salary Slip
	frappe.permissions.can_export("Salary Slip", raise_exception=True)

	filters = frappe.parse_json(filters) or {}
	# Endpoint này chỉ phục vụ bảng lương. `frappe.query_report` là MỘT instance dùng chung cho mọi
	# query report nên nút Export của báo cáo khác rất dễ gọi nhầm vào đây — nói thẳng ra thay vì
	# trả một file rỗng khó hiểu (đúng bẫy đã dính với bảng chấm công, 2026-08-03).
	if not (cint(filters.get("month")) and cint(filters.get("year"))):
		frappe.throw(_("Thiếu tháng/năm: đường xuất Excel này chỉ dùng cho báo cáo Bảng lương MVL."))

	columns, data = execute(filters)

	visible = frappe.parse_json(visible_idx) if visible_idx else None
	if visible:
		keep = set(visible)
		data = [row for i, row in enumerate(data) if i in keep]

	stream = BytesIO()
	signatures = {"prepared_by": prepared_by, "approved_by": approved_by}
	build_workbook(columns, data, filters, signatures).save(stream)

	filters = frappe._dict(filters)
	name = f"Bang luong {cint(filters.month):02d}-{cint(filters.year)}"
	provide_binary_file(name, "xlsx", stream.getvalue())
